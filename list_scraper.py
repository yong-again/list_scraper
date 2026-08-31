"""Listhammer.info AoS 아미 리스트 크롤러 (Playwright 기반).

각 대회 결과 행의 "아미 리스트 보기" 버튼을 클릭해 모달(.p-dialog)에 뜨는
상세 아미 리스트(.army-list)를 파싱해 유닛명/포인트를 수집한다.
페이지네이션이 있으면 마지막 페이지까지 순회한다.

사용 예:
    python list_scraper.py                          # 기본: Kharadron Overlords
    python list_scraper.py --faction "Stormcast Eternals" -o stormcast.csv
    python list_scraper.py --no-headless            # 브라우저 창 표시
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from urllib.parse import quote_plus

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from playwright.sync_api import (
    Error as PlaywrightError,
    Locator,
    Page,
    TimeoutError as PlaywrightTimeoutError,
    sync_playwright,
)

BASE_URL = "https://listhammer.info/aos"
# 2026-08 사이트 개편: 예전에는 행 안에 인라인으로 펼쳐졌지만(.p-datatable-row-toggle-button),
# 지금은 각 행의 "아미 리스트 보기" 버튼(aria-label="View {player}'s army list")을 누르면
# 모달(.p-dialog)이 뜨는 방식으로 바뀌었다.
TOGGLE_BUTTON = '.p-datatable-tbody button[aria-label*="army list" i]'
DIALOG = ".p-dialog"
DIALOG_CLOSE = ".p-dialog-close-button"
ARMY_LIST = ".army-list"
LIST_TITLE = ".text-lg.font-bold"

# .army-list 자식 div들을 브라우저 안에서 한 번에 구조화해서 가져오는 JS.
# (요소마다 locator 왕복을 하면 매우 느리므로 evaluate 한 방으로 처리)
EXTRACT_JS = """
(el) => Array.from(el.querySelectorAll(':scope > div')).map(div => {
    const nameSpan = div.querySelector('span.font-semibold');
    const ptsSpan = div.querySelector('span.text-muted-color, span.text-gray-400');
    return {
        cls: div.getAttribute('class') || '',
        text: div.textContent.trim(),
        unitName: nameSpan ? nameSpan.textContent.trim() : null,
        points: (nameSpan && ptsSpan) ? ptsSpan.textContent.trim() : null,
    };
})
"""


def clean_points(pts: str | None) -> int | None:
    """'260', '130 pts', '20 Points' 같은 표기를 정수로 정규화."""
    if not pts:
        return None
    m = re.search(r"\d+", pts)
    return int(m.group()) if m else None


def trailing_points(text: str | None) -> int | None:
    """유닛 줄 끝에 붙은 포인트를 추출 ('Kragnos, the End of Empires580' → 580).

    현재 listhammer DOM은 포인트를 별도 span이 아니라 유닛명 뒤 텍스트로 붙인다.
    유닛명은 숫자로 끝나지 않으므로 줄 맨 끝의 숫자를 포인트로 본다.
    """
    if not text:
        return None
    m = re.search(r"(\d+)\s*(?:pts?|points?)?\s*$", text.strip(), re.IGNORECASE)
    return int(m.group(1)) if m else None


def parse_army_list(entries: list[dict]) -> dict:
    """EXTRACT_JS 결과(줄 단위 dict 목록)를 리스트 제목/메타/유닛 목록으로 변환."""
    title = None
    meta: dict[str, str] = {}
    units: list[dict] = []
    section = ""  # 현재 연대(General's Regiment, Regiment 1, ...)

    for e in entries:
        cls, text = e["cls"], e["text"]
        if not text or set(text) <= {"-"}:  # 빈 줄, '-----' 구분선 스킵
            continue

        if "text-lg" in cls and "font-bold" in cls:
            title = text
        elif "tracking-wide" in cls:  # 연대/섹션 헤더 (font-bold tracking-wide)
            section = text
        elif e["unitName"]:  # 유닛명 span을 가진 행 (포인트는 span 또는 줄 끝 텍스트)
            name = e["unitName"]
            # 포인트: 예전 형식은 별도 span(e["points"]), 현재 형식은 유닛명 뒤 텍스트에 붙음
            pts_int = clean_points(e["points"]) if e["points"] else trailing_points(text)
            # 일부 빌더 형식: 강화/아티팩트가 "~ " 접두사의 별도 행으로 나옴 → 직전 유닛의 노트로
            if name.startswith("~"):
                label = name.lstrip("~ ").strip()
                note = f"{label} ({pts_int} pts)" if pts_int else label
                if units:
                    units[-1]["notes"].append(note)
                continue
            if name.startswith("Faction Terrain:") or section == "Faction Terrain":
                terrain = name.removeprefix("Faction Terrain:").strip()
                meta["faction_terrain"] = f"{terrain} ({pts_int} pts)" if pts_int else terrain
            # 첫 연대 헤더 이전의 span 라인은 유닛이 아니라 팩션/배틀 포메이션 헤더
            # (GW 앱 형식: "Order | Kharadron Overlords | ...", "Pioneers and Scavengers 10 Points")
            elif not section:
                value = f"{name} ({pts_int})" if pts_int else name
                if "|" in name:  # span 형태의 팩션 줄이 가장 신뢰도 높음 → 덮어씀
                    meta["faction_line"] = value
                else:
                    meta.setdefault("battle_formation", value)
            else:
                units.append(
                    {"section": section, "unit": name, "points": pts_int, "notes": []}
                )
        # 직전 유닛의 옵션/강화 줄 — 예전 형식은 text-gray-400, 현재 형식은 text-muted-color
        # (• General, • Reinforced, - Options:, • 아티팩트 …). 단 text-sm은 구분선/제작툴 푸터라 제외.
        elif ("text-gray-400" in cls or "text-muted-color" in cls) and "text-sm" not in cls:
            if units:
                units[-1]["notes"].append(text.lstrip("•*-–— ").strip())
        elif cls == "":  # 클래스 없는 메타데이터 줄 (Points:, Drops:, Battle Tactics:, 팩션)
            key, _, val = text.partition(":")
            if val:
                key = key.strip().lower().replace(" ", "_")
                if key.startswith("battle_tactic"):  # "Battle Tactic(s) Cards" 등 변형 통일
                    key = "battle_tactics"
                meta[key] = val.strip()
            elif section == "Faction Terrain":
                # span 없이 섹션 헤더 다음 줄에 이름만 오는 형식 (예: "Shrine Luminor")
                meta.setdefault("faction_terrain", text)
            elif section:
                # 일부 빌더 형식: 유닛이 span 없이 "Arkanaut Company (180) (ALWAYS VETERAN)"처럼 표기됨
                m = re.match(r"^(.+?)\s*\((\d+)\)\s*(.*)$", text)
                if m:
                    note = m.group(3).strip("() ")
                    units.append({"section": section, "unit": m.group(1).strip(),
                                  "points": int(m.group(2)), "notes": [note] if note else []})
            elif re.fullmatch(r"\d+\s*/\s*\d+\s*(?:pts|points)?", text, re.IGNORECASE):
                meta.setdefault("points", text)  # "2000/2000 pts" 단독 줄 형식
            elif "faction_line" not in meta:
                meta["faction_line"] = text
            else:  # 팩션명 다음 줄들 (배틀 포메이션, General's Handbook 버전 등)
                meta.setdefault("subtitle_lines", []).append(text)

    # GW 앱 형식은 "Points:" 줄이 없고 제목에 "1990/2000 pts"처럼 포함됨
    if "points" not in meta and title:
        m = re.search(r"(\d+\s*/\s*\d+)\s*(?:pts|points)", title, re.IGNORECASE)
        if m:
            meta["points"] = m.group(1).replace(" ", "")

    return {"title": title, "meta": meta, "units": units}


def scrape_row(page: Page, button: Locator, timeout_ms: int,
               dump_path: Path | None = None, retries: int = 1) -> dict | None:
    """"아미 리스트 보기" 버튼을 클릭해 모달(.p-dialog)에 뜨는 .army-list를 파싱. 실패 시 None.

    모달 내용은 행마다 서버에서 따로 불러오는데 응답이 느릴 때가 있어, 타임아웃이 나면
    (열린 채로 멈춰 있는 모달을 정리한 뒤) 최대 `retries`번 다시 시도한다.

    dump_path가 주어지면 모달에 뜬 .army-list의 원본 HTML을 그대로 저장한다
    (빌더별 DOM 구조 확인·파서 디버깅용).
    """
    # PrimeVue Dialog는 페이지에 하나뿐인 모달을 재사용한다 (Teleport 방식).
    dialog = page.locator(DIALOG)
    army = dialog.locator(ARMY_LIST)

    for attempt in range(retries + 1):
        button.scroll_into_view_if_needed()
        button.click()
        try:
            # 명시적 대기: 모달 안 내용이 실제로 채워질 때까지
            army.locator(f"{LIST_TITLE}, span.font-semibold").first.wait_for(
                state="attached", timeout=timeout_ms
            )
            break
        except PlaywrightTimeoutError:
            # 다음 시도 전에, 열린 채로 멈춰 있을 수 있는 모달을 정리
            try:
                dialog.locator(DIALOG_CLOSE).click(timeout=2000)
                dialog.wait_for(state="detached", timeout=5000)
            except PlaywrightError:
                pass
            if attempt == retries:
                raise
            print(f"    (재시도 {attempt + 1}/{retries}) 로딩 시간 초과 — 다시 시도합니다")

    if dump_path is not None:
        dump_path.write_text(army.evaluate("el => el.outerHTML"), encoding="utf-8")
    parsed = parse_army_list(army.evaluate(EXTRACT_JS))

    # 모달을 완전히 닫을 때까지 대기 (다음 행 클릭 시 이전 모달과 겹치지 않도록)
    try:
        dialog.locator(DIALOG_CLOSE).click()
        dialog.wait_for(state="detached", timeout=timeout_ms)
    except PlaywrightError:
        pass  # 닫기 실패는 치명적이지 않음

    return parsed


NOT_A_FORMATION = re.compile(r"general.?s handbook|created with|game version|grand alliance", re.IGNORECASE)


def battle_formation(row_meta: dict, meta: dict, valid: set[str] | None = None) -> str:
    """배틀 포메이션 결정: 테이블 셀 → 상세 메타 → 팩션 라인 → 부제 줄 순으로 폴백.

    valid(wahapedia 기준 유효 포메이션 목록)가 주어지면 후보 줄에서 유효한
    이름이 포함된 것만 인정한다 — 빌더마다 헤더 줄 구성이 달라 오탐이 잦기 때문.
    """
    candidates = []
    if "|" in row_meta["faction"]:
        candidates.append(row_meta["faction"].split("|")[-1])
    candidates.append(meta.get("battle_formation", ""))
    fl = meta.get("faction_line", "")
    if "|" in fl:
        candidates.append(fl.split("|")[-1])
    candidates += meta.get("subtitle_lines", [])
    # "Pioneers and Scavengers (10 Points)" 같은 포인트 표기 제거
    candidates = [re.sub(r"\s*\(\d+[^)]*\)$", "", c.strip()).strip()
                  for c in candidates if c.strip() and not NOT_A_FORMATION.search(c)]
    if valid:
        for c in candidates:
            for f in valid:
                if f.lower() in c.lower():
                    return f
        return ""
    return candidates[0] if candidates else ""


def load_valid_formations(faction: str) -> set[str]:
    """wahapedia 팩션 룰에서 배틀 포메이션 이름 목록 추출."""
    path = FACTION_RULES_DIR / f"{faction_slug(faction)}.json"
    if not path.exists():
        return set()
    out = set()
    for x in json.loads(path.read_text()):
        if x["category"] == "army_rules":
            m = re.match(r"Battle Formations > (.+)", x["section"])
            if m:
                out.add(m.group(1).strip())
    return out


def row_metadata(row: Locator) -> dict:
    """테이블 행 자체의 메타데이터(플레이어/이벤트/날짜/전적) 추출."""
    cells = [c.strip() for c in row.locator("td").all_inner_texts()]
    # 컬럼: [토글, Name, Faction\n서브팩션, Event\n(n players), 날짜, Result]
    get = lambda i: cells[i] if i < len(cells) else ""
    return {
        "player": get(1),
        "faction": get(2).replace("\n", " | "),
        "event": get(3).replace("\n", " "),
        "date": get(4),
        "result": get(5),
    }


def go_next_page(page: Page) -> bool:
    """다음 페이지 버튼 클릭. 마지막 페이지(disabled)면 False."""
    next_btn = page.locator(".p-paginator button:has(.pi-chevron-right)").first
    if next_btn.count() == 0 or next_btn.is_disabled():
        return False
    first_row = page.locator(".p-datatable-tbody > tr").first
    next_btn.click()
    try:
        first_row.wait_for(state="detached", timeout=15000)  # 기존 행 교체 대기
    except PlaywrightTimeoutError:
        pass
    page.wait_for_selector(TOGGLE_BUTTON, timeout=15000)
    return True


def scrape(url: str, headless: bool, timeout_ms: int, max_pages: int | None,
           faction: str = "", dump_html: Path | None = None, retries: int = 1) -> pd.DataFrame:
    records: list[dict] = []
    valid_formations = load_valid_formations(faction) if faction else set()
    if dump_html is not None:
        dump_html.mkdir(parents=True, exist_ok=True)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=headless)
        page = browser.new_page()
        print(f"접속 중: {url}")
        page.goto(url, wait_until="networkidle", timeout=60000)

        try:
            page.wait_for_selector(TOGGLE_BUTTON, timeout=timeout_ms)
        except PlaywrightTimeoutError:
            print("확장 버튼을 찾지 못했습니다. 팩션명/URL을 확인하세요.", file=sys.stderr)
            browser.close()
            return pd.DataFrame()

        page_no = 1
        while True:
            buttons = page.locator(TOGGLE_BUTTON)
            n = buttons.count()
            print(f"[페이지 {page_no}] 행 {n}개 발견")

            for i in range(n):
                button = buttons.nth(i)
                row_meta = row_metadata(button.locator("xpath=ancestor::tr[1]"))
                dump_path = None
                if dump_html is not None:
                    safe = re.sub(r"[^\w.-]+", "_", f"{row_meta['player']}_{page_no}_{i}") or f"row_{page_no}_{i}"
                    dump_path = dump_html / f"{safe}.html"
                try:
                    parsed = scrape_row(page, button, timeout_ms, dump_path, retries)
                except PlaywrightTimeoutError:
                    print(f"  ({i + 1}/{n}) {row_meta['player']}: 상세 내용 로딩 시간 초과 "
                          f"({retries + 1}번 시도) — 건너뜀")
                    continue
                except PlaywrightError as exc:
                    print(f"  ({i + 1}/{n}) {row_meta['player']}: 오류 발생 — {exc}")
                    continue

                if not parsed or not parsed["units"]:
                    print(f"  ({i + 1}/{n}) {row_meta['player']}: 유닛 데이터 없음 — 건너뜀")
                    continue

                for u in parsed["units"]:
                    records.append({
                        **row_meta,
                        "battle_formation": battle_formation(row_meta, parsed["meta"], valid_formations),
                        "list_title": parsed["title"],
                        "total_points": parsed["meta"].get("points", ""),
                        "drops": parsed["meta"].get("drops", ""),
                        "battle_tactics": parsed["meta"].get("battle_tactics", ""),
                        "faction_terrain": parsed["meta"].get("faction_terrain", ""),
                        "regiment": u["section"],
                        "unit": u["unit"],
                        "points": u["points"],
                        "notes": "; ".join(u["notes"]),
                    })
                print(f"  ({i + 1}/{n}) {row_meta['player']}: 유닛 {len(parsed['units'])}개 수집")

            if max_pages and page_no >= max_pages:
                break
            if not go_next_page(page):
                break
            page_no += 1

        browser.close()

    return pd.DataFrame(records)


# ---------------------------------------------------------------------------
# Excel 리포트 (아미 리스트 / 팩션 룰 번역 / 통계 분석)
# ---------------------------------------------------------------------------

FACTION_RULES_DIR = Path("wahapedia_factions")
WARSCROLL_DIR = Path("warscolls")
TRANSLATION_DIR = Path("translations")

# 워스크롤 키워드 → 역할 분류 우선순위 (앞쪽이 우선)
ROLE_PRECEDENCE = ["HERO", "MONSTER", "WAR MACHINE", "CAVALRY", "BEAST",
                   "INFANTRY", "MANIFESTATION", "FACTION TERRAIN"]

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=13)
WRAP = Alignment(wrap_text=True, vertical="top")


def faction_slug(faction: str) -> str:
    return re.sub(r"[’']", "", faction).lower().replace(" ", "-")


def norm_name(name: str) -> str:
    """유닛명 매칭용 정규화 — 쉼표/따옴표 등 표기 차이 무시 (예: 'Brokk Grungsson, Lord-Magnate')."""
    return re.sub(r"\s+", " ", re.sub(r"[,’'.]", "", name.lower())).strip()


def load_rules_sheet_rows(faction: str) -> list[dict]:
    """wahapedia 팩션 룰(army_rules) + 한국어 번역을 행 목록으로 반환."""
    slug = faction_slug(faction)
    rules_path = FACTION_RULES_DIR / f"{slug}.json"
    if not rules_path.exists():
        return [{"섹션": "-", "원문": f"팩션 룰 파일 없음: {rules_path}", "한국어 번역": ""}]

    rules = [x for x in json.loads(rules_path.read_text()) if x["category"] == "army_rules"]

    trans_path = TRANSLATION_DIR / f"{slug}.ko.json"
    translations = []
    if trans_path.exists():
        translations = json.loads(trans_path.read_text())["entries"]

    rows = []
    for i, r in enumerate(rules):
        ko = ""
        if i < len(translations) and translations[i]["section"] == r["section"]:
            ko = translations[i]["ko"]
        elif not translations:
            ko = f"(번역 파일 없음: {trans_path})"
        rows.append({"섹션": r["section"] or "소개", "원문": r["text"], "한국어 번역": ko})
    return rows


def load_unit_roles(faction: str) -> dict[str, str]:
    """워스크롤 키워드를 기반으로 유닛명(소문자) → 역할 매핑 생성."""
    path = WARSCROLL_DIR / f"{faction_slug(faction)}.json"
    if not path.exists():
        return {}
    roles = {}
    for ws in json.loads(path.read_text()):
        keywords = set(ws.get("keywords", []))
        role = next((r for r in ROLE_PRECEDENCE if r in keywords), "기타")
        roles[norm_name(ws["name"])] = role
    return roles


def load_unit_points(faction: str) -> dict[str, int]:
    """워스크롤 기준 유닛 포인트 매핑 (Battle Profiles로 갱신된 현재 값)."""
    path = WARSCROLL_DIR / f"{faction_slug(faction)}.json"
    if not path.exists():
        return {}
    points = {}
    for ws in json.loads(path.read_text()):
        if isinstance(ws.get("points"), int):
            points[norm_name(ws["name"])] = ws["points"]
    return points


ENHANCEMENT_EXCLUDE = re.compile(r"^(General|Reinforced|Options\b|\d+\s*x?\s)", re.IGNORECASE)


def enrich_units_df(df: pd.DataFrame, roles: dict[str, str],
                    unit_points: dict[str, int] | None = None) -> pd.DataFrame:
    """통계 계산용 파생 컬럼(list_id/base_unit/role/unit_points/formation/승패) 추가."""
    d = df.copy()
    d["list_id"] = d[["player", "event", "date"]].astype(str).agg("|".join, axis=1)
    d["base_unit"] = (d["unit"]
                      .str.replace(r"^\d+\s*x\s+", "", regex=True, flags=re.IGNORECASE)
                      .str.replace(r"^\d+(?:st|nd|rd|th)\s*model\s*-\s*", "", regex=True, flags=re.IGNORECASE)
                      .str.replace(r"^SoG\s+", "Scourge of Ghyran ", regex=True)  # 일부 빌더의 약어 표기
                      .str.strip())

    def canon(name: str) -> str:
        """복수형 표기("Wrekkazs")가 워스크롤 이름과 어긋나면 단수형으로 통일."""
        key = norm_name(name)
        if key not in roles and key.endswith("s") and key[:-1] in roles:
            return name.rstrip("sS")
        return name

    d["base_unit"] = d["base_unit"].map(canon)
    d["role"] = d["base_unit"].map(lambda n: roles.get(norm_name(n), "기타"))
    # 워스크롤 기준 포인트 (없는 유닛은 NaN → 집계 시 관측값 최빈치로 폴백)
    d["unit_points"] = d["base_unit"].map(lambda n: (unit_points or {}).get(norm_name(n)))
    if "battle_formation" in d.columns:
        d["formation"] = d["battle_formation"].fillna("").replace("", "(불명)")
    else:
        d["formation"] = d["faction"].str.split("|").str[-1].str.strip()
    res = d["result"].astype(str).str.extract(r"(\d+)\s*-\s*(\d+)")
    d["wins"], d["losses"] = res[0].astype(float), res[1].astype(float)
    return d


def enhancement_counts(d: pd.DataFrame) -> pd.DataFrame:
    """노트 컬럼에서 강화 채용 횟수 집계 (장비 옵션/General/Reinforced 제외)."""
    notes = (d["notes"].fillna("").str.split(";").explode().str.strip().str.lstrip("•*-– "))
    notes = notes[(notes != "") & ~notes.str.match(ENHANCEMENT_EXCLUDE)]
    notes = notes[~notes.str.isupper()]  # "ALWAYS VETERAN" 같은 전부 대문자 장비 표기 제외
    # 빌더별 포인트 표기 통일: "(10 pts)", "- (10) Points", "10 pts" 등 제거
    notes = notes.str.replace(r"\s*[-–]?\s*\(\d+\)\s*(?:pts?|points?)?\s*$", "", regex=True, flags=re.IGNORECASE)
    notes = notes.str.replace(r"\s*\(?\d+\s*(?:pts?|points?)\)?\s*$", "", regex=True, flags=re.IGNORECASE).str.strip()
    # 하이픈/대소문자 표기 차이 통합 (예: "Celestium Burst-grenade" = "Celestium Burst Grenade")
    tmp = pd.DataFrame({"raw": notes})
    tmp["canon"] = tmp["raw"].str.lower().str.replace("-", " ").str.replace(r"\s+", " ", regex=True)
    return (tmp.groupby("canon")
            .agg(강화=("raw", lambda s: s.value_counts().index[0]), 채용_횟수=("raw", "size"))
            .sort_values("채용_횟수", ascending=False).reset_index(drop=True))


def build_stats_tables(df: pd.DataFrame, roles: dict[str, str],
                       unit_points: dict[str, int] | None = None) -> list[tuple[str, pd.DataFrame]]:
    """통계 분석 시트에 들어갈 (제목, 표) 목록 생성."""
    d = enrich_units_df(df, roles, unit_points)
    lists = d.drop_duplicates("list_id")
    n_lists = len(lists)
    tables: list[tuple[str, pd.DataFrame]] = []

    # 1. 개요
    overview = pd.DataFrame([
        ("분석 리스트 수", n_lists),
        ("참가 이벤트 수", lists["event"].nunique()),
        ("평균 전적 (승)", round(lists["wins"].mean(), 2)),
        ("평균 드랍 수", round(pd.to_numeric(lists["drops"].astype(str).str.extract(r"(\d+)")[0], errors="coerce").mean(), 2)),
        ("팩션 터레인 채용률", f"{lists['faction_terrain'].replace('', pd.NA).notna().mean():.0%}"),
    ], columns=["항목", "값"])
    tables.append(("개요", overview))

    # 2. 배틀 포메이션
    formation = (lists.groupby("formation")
                 .agg(리스트_수=("list_id", "count"), 평균_승수=("wins", "mean"), 평균_패수=("losses", "mean"))
                 .round(2).sort_values("리스트_수", ascending=False).reset_index()
                 .rename(columns={"formation": "배틀 포메이션"}))
    formation.insert(0, "순위", range(1, len(formation) + 1))
    tables.append(("가장 많이 쓰인 배틀 포메이션", formation))

    # 2-1. 팩션 터레인
    terr = lists["faction_terrain"].fillna("").astype(str).str.strip()
    terr_names = terr.mask(terr == "", "(미채용)").str.replace(r"\s*\(\d+[^)]*\)$", "", regex=True)
    terrain = terr_names.value_counts().reset_index()
    terrain.columns = ["팩션 터레인", "리스트_수"]
    terrain["비율"] = (terrain["리스트_수"] / n_lists).map("{:.0%}".format)
    tables.append(("팩션 터레인", terrain))

    # 3. 유닛 사용 TOP 10 (전체) — 포인트는 워스크롤 기준값 (없으면 관측 최빈치)
    def unit_table(sub: pd.DataFrame, top: int) -> pd.DataFrame:
        g = (sub.groupby("base_unit")
             .agg(채용_리스트_수=("list_id", "nunique"), 총_채용_횟수=("list_id", "count"),
                  포인트=("unit_points", "first"),
                  관측_포인트=("points", lambda s: s.mode().iat[0]))
             .sort_values(["채용_리스트_수", "총_채용_횟수"], ascending=False).head(top).reset_index()
             .rename(columns={"base_unit": "유닛"}))
        g["채용률"] = (g["채용_리스트_수"] / n_lists).map("{:.0%}".format)
        g["포인트"] = g["포인트"].fillna(g["관측_포인트"]).astype(int)
        g.insert(0, "순위", range(1, len(g) + 1))
        return g[["순위", "유닛", "채용_리스트_수", "채용률", "총_채용_횟수", "포인트"]]

    tables.append(("가장 많이 쓰인 유닛 TOP 10 (전체)", unit_table(d, 10)))

    # 4. 역할(구성)별 유닛 TOP 5
    for role in [r for r in ROLE_PRECEDENCE + ["기타"] if r in set(d["role"])]:
        sub = d[d["role"] == role]
        tables.append((f"가장 많이 쓰인 유닛 TOP 5 — {role}", unit_table(sub, 5)))

    # 5. 강화(Enhancements) TOP 10 — 노트에서 장비 옵션/General/Reinforced 제외
    enh = enhancement_counts(d).head(10)
    enh.insert(0, "순위", range(1, len(enh) + 1))
    tables.append(("가장 많이 쓰인 강화 TOP 10", enh))

    # 6. 장군(General) 선택
    gen = d[d["notes"].fillna("").str.contains(r"\bGeneral\b")]
    gen_t = gen["base_unit"].value_counts().reset_index()
    gen_t.columns = ["장군으로 지정된 유닛", "리스트 수"]
    gen_t["비율"] = (gen_t["리스트 수"] / n_lists).map("{:.0%}".format)
    tables.append(("장군(General) 선택", gen_t))

    return tables


def style_header(ws, row: int, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT


def write_excel(df: pd.DataFrame, faction: str, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        # 시트 1: 아미 리스트 원본 데이터
        df.to_excel(writer, sheet_name="아미 리스트", index=False)
        ws = writer.sheets["아미 리스트"]
        style_header(ws, 1, len(df.columns))
        ws.freeze_panes = "A2"
        widths = {"player": 18, "faction": 32, "event": 34, "list_title": 30, "unit": 34,
                  "battle_formation": 26,
                  "battle_tactics": 26, "faction_terrain": 24, "regiment": 18, "notes": 45}
        for i, col in enumerate(df.columns, 1):
            ws.column_dimensions[get_column_letter(i)].width = widths.get(col, 12)

        # 시트 2: 팩션 룰 번역
        rules_df = pd.DataFrame(load_rules_sheet_rows(faction))
        rules_df.to_excel(writer, sheet_name="팩션 룰(번역)", index=False)
        ws = writer.sheets["팩션 룰(번역)"]
        style_header(ws, 1, 3)
        ws.freeze_panes = "A2"
        for col, w in zip("ABC", (34, 70, 70)):
            ws.column_dimensions[col].width = w
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = WRAP

        # 시트 3: 통계 분석
        ws = writer.book.create_sheet("통계 분석")
        r = 1
        for title, table in build_stats_tables(df, load_unit_roles(faction),
                                               load_unit_points(faction)):
            ws.cell(row=r, column=1, value=title).font = TITLE_FONT
            r += 1
            for c, col in enumerate(table.columns, 1):
                ws.cell(row=r, column=c, value=col)
            style_header(ws, r, len(table.columns))
            r += 1
            for rec in table.itertuples(index=False):
                for c, v in enumerate(rec, 1):
                    ws.cell(row=r, column=c, value=v)
                r += 1
            r += 2  # 표 사이 간격
        for col, w in zip("ABCDEF", (44, 22, 16, 12, 14, 12)):
            ws.column_dimensions[col].width = w


def main() -> None:
    ap = argparse.ArgumentParser(description="listhammer.info AoS 아미 리스트 크롤러")
    ap.add_argument("--faction", default="Kharadron Overlords", help="팩션명 (기본: Kharadron Overlords)")
    ap.add_argument("--url", help="전체 URL 직접 지정 (--faction 무시)")
    ap.add_argument("--csv", help="CSV도 함께 저장할 경로 (선택)")
    ap.add_argument("--no-headless", action="store_true", help="브라우저 창을 띄워서 실행")
    ap.add_argument("--timeout", type=int, default=15000, help="명시적 대기 타임아웃 (ms)")
    ap.add_argument("--retries", type=int, default=1,
                    help="행별 로딩 타임아웃 시 재시도 횟수 (기본: 1번 재시도, 총 2번 시도)")
    ap.add_argument("--max-pages", type=int, default=None, help="최대 크롤링 페이지 수")
    ap.add_argument("--dump-html", metavar="DIR",
                    help="확장된 army-list 원본 HTML을 DIR에 저장 (파서 디버깅용)")
    args = ap.parse_args()

    url = args.url or f"{BASE_URL}?faction={quote_plus(args.faction)}"
    df = scrape(url, headless=not args.no_headless, timeout_ms=args.timeout,
                max_pages=args.max_pages, faction=args.faction,
                dump_html=Path(args.dump_html) if args.dump_html else None,
                retries=args.retries)

    if df.empty:
        print("수집된 데이터가 없습니다.")
        sys.exit(1)

    xlsx_path = Path("army_list") / f"{args.faction.replace(' ', '_')}_army_list.xlsx"
    write_excel(df, args.faction, xlsx_path)
    print(f"\n총 {len(df)}개 유닛 행 저장 완료 → {xlsx_path}")
    print("시트: 아미 리스트 / 팩션 룰(번역) / 통계 분석")

    if args.csv:
        df.to_csv(args.csv, index=False, encoding="utf-8-sig")
        print(f"CSV 저장 → {args.csv}")


if __name__ == "__main__":
    main()
